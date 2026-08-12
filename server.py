"""
server.py — Flask backend cho MNT_FB
Toàn bộ data từ SQLite local, không cần Google Sheets API.
"""

import os
import sys
import json
import time
import subprocess
from datetime import datetime
from pathlib import Path
from flask import Flask, jsonify, render_template, request, send_from_directory

BASE_DIR = Path(__file__).parent
sys.path.insert(0, str(BASE_DIR))
os.chdir(str(BASE_DIR))

from config import PORT, LOG_DIR, MEDIA_DIR
import db
from db import (
    init_db,
    # accounts
    get_accounts, get_account_by_name, upsert_account, delete_account, update_account_field,
    reorder_accounts, insert_account_at, import_accounts,
    # pages
    get_pages, get_page_by_name, upsert_page, delete_page, reorder_pages, import_pages,
    # content
    get_content, get_content_by_code, upsert_content, delete_content, reorder_content,
    # uid groups
    get_uid_groups_by_code, get_all_uid_groups, upsert_uid_group, delete_uid_group,
    reorder_uid_groups, import_uid_groups,
    # schedules
    get_schedules, update_schedule_status, bulk_set_schedule_status,
    replace_schedules, update_schedule_field,
    # comment posts
    get_comment_posts, them_comment_posts, update_comment_post_field,
    delete_comment_post, xoa_het_comment_posts,
    # loại đăng
    LOAI_DANG_OPTIONS, LOAI_LICH_MAP, la_loai_comment, la_loai_hon_hop,
    TI_LE_COMMENT_MAC_DINH, accounts_theo_lich,
    # settings
    get_setting, set_setting, get_all_settings,
)
from utils import logger

# Init DB on startup
init_db()

app = Flask(__name__, template_folder="templates", static_folder="static")
app.json.ensure_ascii = False
# Không cache file tĩnh (js/css) — luôn nạp bản mới sau khi sửa, khỏi phải Ctrl+F5
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0
# Tự nạp lại template khi sửa index.html — khỏi phải restart server
app.config["TEMPLATES_AUTO_RELOAD"] = True

# ── CHỈ CHẠY TẠI MÁY ─────────────────────────────────────────────
# Đã bỏ hẳn điều khiển từ xa (Tailscale) cùng toàn bộ lớp đăng nhập/mật khẩu.
# Server lắng nghe 127.0.0.1 (xem _serve) nên chỉ máy này gọi được — không còn
# máy nào trong mạng LAN chạm tới cổng 8080, và cũng không cần lớp xác thực nào.

# Boot ID đổi mỗi lần server khởi động — client dùng để phát hiện restart và tự reload.
APP_BOOT_ID = str(time.time())


@app.route("/api/ping")
def api_ping():
    return jsonify({"ok": True, "boot": APP_BOOT_ID})

# ── Serve media files ─────────────────────────────────────────
@app.route("/media/<path:filename>")
def serve_media(filename):
    return send_from_directory(str(MEDIA_DIR), filename)

@app.route("/data/media/<path:filename>")
def serve_media_alt(filename):
    return send_from_directory(str(MEDIA_DIR), filename)


# ── Runner management ─────────────────────────────────────────
RUNNER_CFG = {
    "homestay": {"sheet": "Chéo Homestay", "pid_file": ".runner_homestay.pid",
                 "log": str(LOG_DIR / "autopost_homestay.log")},
    "thue":     {"sheet": "Chéo Thuê",     "pid_file": ".runner_thue.pid",
                 "log": str(LOG_DIR / "autopost_thue.log")},
    "ban":      {"sheet": "Chéo Bán",      "pid_file": ".runner_ban.pid",
                 "log": str(LOG_DIR / "autopost_ban.log")},
    "page":     {"sheet": "Đăng bài Page", "pid_file": ".runner_page.pid",
                 "log": str(LOG_DIR / "autopost_page.log")},
    "nuoi":     {"sheet": "Nuôi nick",     "pid_file": ".runner_nuoi.pid",
                 "log": str(LOG_DIR / "autopost_nuoi.log")},
}
RUNNER_LOAI_MAP = {
    "homestay": "homestay",
    "thue":     "thue",
    "ban":      "ban",
    "page":     "page",
    "nuoi":     "nuoi",
}
STAGGER = {"homestay": 0, "thue": 8, "ban": 16, "page": 24, "nuoi": 32}


def _runner_pid(loai):
    pf = BASE_DIR / RUNNER_CFG[loai]["pid_file"]
    try:
        return int(pf.read_text().strip()) if pf.exists() else None
    except Exception:
        return None


def _pid_alive(pid: int) -> bool:
    """
    Kiểm tra PID còn sống — KHÔNG dùng os.kill(pid, 0).
    Trên Windows os.kill(pid,0) gọi GenerateConsoleCtrlEvent (gửi Ctrl+C),
    raise lỗi khi server chạy dưới pythonw (không console) → false dương tính.
    """
    if sys.platform == "win32":
        import ctypes
        PROCESS_QUERY_LIMITED_INFORMATION = 0x1000
        STILL_ACTIVE = 259
        k32 = ctypes.windll.kernel32
        h = k32.OpenProcess(PROCESS_QUERY_LIMITED_INFORMATION, False, pid)
        if not h:
            return False
        try:
            code = ctypes.c_ulong()
            ok = k32.GetExitCodeProcess(h, ctypes.byref(code))
            return bool(ok) and code.value == STILL_ACTIVE
        finally:
            k32.CloseHandle(h)
    try:
        os.kill(pid, 0)
        return True
    except OSError:
        return False


def _find_python_pids(*needles) -> list:
    """
    Tìm PID tiến trình python có dòng lệnh chứa TẤT CẢ chuỗi trong `needles`.

    Dùng PowerShell/CIM thay cho `wmic`: Microsoft đang gỡ dần wmic khỏi Windows
    11. Mất nó thì không diệt được scheduler mồ côi, dẫn tới hai runner cùng chạy
    trên một profile Chrome — đúng thứ làm hỏng phiên đăng nhập.
    """
    if sys.platform != "win32":
        return []
    dk = " -and ".join(f"$_.CommandLine -like '*{n}*'" for n in needles)
    ps = ("Get-CimInstance Win32_Process "
          "-Filter \"Name='python.exe' OR Name='pythonw.exe'\" | "
          f"Where-Object {{ {dk} }} | ForEach-Object {{ $_.ProcessId }}")
    try:
        r = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps],
            capture_output=True, text=True, timeout=15,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
        return [int(x) for x in r.stdout.split() if x.strip().isdigit()]
    except Exception as e:
        logger.warning(f"Không dò được PID python: {e}")
        return []


def _kill_pids(pids) -> list:
    """taskkill /F /T từng PID, trả về danh sách đã diệt."""
    da_diet = []
    for pid in pids:
        try:
            subprocess.run(["taskkill", "/F", "/T", "/PID", str(pid)],
                           capture_output=True,
                           creationflags=subprocess.CREATE_NO_WINDOW)
            da_diet.append(int(pid))
        except Exception:
            pass
    return da_diet


def _runner_running(loai):
    pid = _runner_pid(loai)
    if not pid:
        return False
    if _pid_alive(pid):
        return True
    (BASE_DIR / RUNNER_CFG[loai]["pid_file"]).unlink(missing_ok=True)
    return False


def _kill_all_runners():
    """Kill tất cả scheduler process — kể cả orphan không có pid file."""
    # Bước 1: kill theo pid file
    for loai, cfg in RUNNER_CFG.items():
        pf = BASE_DIR / cfg["pid_file"]
        if pf.exists():
            try:
                pid = int(pf.read_text().strip())
                if sys.platform == "win32":
                    subprocess.run(["taskkill", "/F", "/T", "/PID", str(pid)],
                                   capture_output=True)
                else:
                    os.kill(pid, 15)
            except Exception:
                pass
            pf.unlink(missing_ok=True)

    # Bước 2: quét mọi scheduler.py mồ côi (không còn pid file)
    for pid in _kill_pids(_find_python_pids("scheduler.py")):
        logger.info(f"  Đã diệt scheduler mồ côi PID {pid}")


def _kill_join_workers():
    """Kill mọi tiến trình join_groups_worker.py đang chạy."""
    for pid in _kill_pids(_find_python_pids("join_groups_worker")):
        logger.info(f"  Đã diệt join worker PID {pid}")


def _shutdown_all():
    """Tắt sạch: server (đang tự thoát) + runner đăng nền + join worker."""
    logger.info("🛑 Đóng app — tắt toàn bộ runner nền...")
    try:
        _kill_all_runners()
        _kill_join_workers()
    except Exception as e:
        logger.warning(f"shutdown lỗi: {e}")


# ═══════════════════════════════════════════════════════════════
# Pages — HTML
# ═══════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/app/shutdown", methods=["POST"])
def api_app_shutdown():
    """Tắt sạch phần mềm: dừng runner + join worker rồi thoát hẳn server.
    Dùng khi muốn giải phóng tài nguyên (bật lại: khởi động máy / RUN_APP)."""
    import threading as _th
    def _bye():
        time.sleep(0.6)          # để response kịp trả về client
        _shutdown_all()
        os._exit(0)              # thoát cả server (và cửa sổ app nếu có)
    _th.Thread(target=_bye, daemon=True).start()
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════
# API — Runner
# ═══════════════════════════════════════════════════════════════

@app.route("/api/run/status")
def run_status():
    return jsonify({loai: {"running": _runner_running(loai)} for loai in RUNNER_CFG})


@app.route("/api/run/<loai>/start", methods=["POST"])
def run_start(loai):
    if loai not in RUNNER_CFG:
        return jsonify({"ok": False, "error": "Loại không hợp lệ"})
    if _runner_running(loai):
        return jsonify({"ok": False, "msg": f"Runner {loai} đang chạy rồi"})
    cfg      = RUNNER_CFG[loai]
    delay    = STAGGER.get(loai, 0)
    headless = (request.json or {}).get("headless", True)
    flags    = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
    env      = {**os.environ,
                "SCHEDULER_LOAI":        loai,
                "SCHEDULER_LOG_FILE":    cfg["log"],
                "SCHEDULER_START_DELAY": str(delay),
                "HEADLESS":              "true" if headless else "false"}
    # Truyền `loai` cả qua dòng lệnh (không chỉ biến môi trường) để lúc cần còn
    # nhận ra runner nào là của loại nào mà diệt đúng cái mồ côi.
    proc  = subprocess.Popen([sys.executable, "-X", "utf8", "scheduler.py", loai],
                             cwd=str(BASE_DIR), creationflags=flags, env=env)
    (BASE_DIR / cfg["pid_file"]).write_text(str(proc.pid))
    return jsonify({"ok": True, "pid": proc.pid})


@app.route("/api/run/<loai>/stop", methods=["POST"])
def run_stop(loai):
    if loai not in RUNNER_CFG:
        return jsonify({"ok": False, "error": "Loại không hợp lệ"})
    killed = []
    # Kill theo pid file
    pid = _runner_pid(loai)
    if pid:
        try:
            if sys.platform == "win32":
                subprocess.run(["taskkill", "/F", "/T", "/PID", str(pid)], capture_output=True)
            else:
                os.kill(pid, 15)
            killed.append(pid)
        except Exception:
            pass
        (BASE_DIR / RUNNER_CFG[loai]["pid_file"]).unlink(missing_ok=True)

    # Quét orphan CÙNG LOẠI (pid file mất/cũ). Nhận ra nhau nhờ `loai` nằm trên
    # dòng lệnh — trước đây lọc theo loai nhưng loai chỉ có trong biến môi
    # trường nên bộ lọc không bao giờ khớp, phần này coi như vô tác dụng.
    # Khớp cụm LIỀN "scheduler.py <loai>" đúng như lúc khởi chạy, thay vì hai
    # chuỗi rời — rời rạc thì một tiến trình python bất kỳ nhắc tới cả hai chữ
    # cũng bị tính là runner.
    killed += _kill_pids(_find_python_pids(f"scheduler.py {RUNNER_LOAI_MAP[loai]}"))
    return jsonify({"ok": True, "killed": killed})


# ═══════════════════════════════════════════════════════════════
# API — Accounts
# ═══════════════════════════════════════════════════════════════

# Cột khai báo INTEGER trong DB. Sửa trực tiếp trên bảng gửi lên chuỗi JSON,
# mà SQLite dùng kiểu động nên nhận chuỗi tuốt — ghi vào thì im lặng, tới lúc
# đọc ra so sánh mới vỡ. Đã gặp thật: ô "Bài đăng tối đa" bị xoá trống thành ''
# khiến gen lịch Page lỗi HTTP 500 ('' > 0 không so sánh được).
COT_SO = {"bai_dang_toi_da", "thoi_gian_nghi", "nuoi_nick", "nuoi_interval",
          "order_idx"}


def ep_kieu_so(field: str, value):
    """Ép giá trị về int nếu `field` là cột số. Rỗng/rác -> 0."""
    if field not in COT_SO:
        return value
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0


# Các trường credential không bao giờ được trả về trong danh sách tài khoản.
# Muốn xem giá trị thật phải gọi /api/accounts/<id>/secrets từ máy local.
SECRET_ACCOUNT_FIELDS = ("password", "xs", "twofa", "pass_khoiphuc", "email_khoiphuc")
SECRET_MASK = "••••••"


def _mask_account(row: dict) -> dict:
    out = dict(row)
    for f in SECRET_ACCOUNT_FIELDS:
        out[f] = SECRET_MASK if (out.get(f) or "").strip() else ""
    return out


@app.route("/api/accounts")
def api_accounts():
    loai = request.args.get("loai")
    rows = get_accounts(loai=loai)
    return jsonify({"ok": True, "data": [_mask_account(r) for r in rows]})


@app.route("/api/accounts/<int:acc_id>/secrets")
def api_account_secrets(acc_id):
    """Trả về credential thật của 1 tài khoản (để sửa trên bảng Tài khoản).

    Không cần chặn theo IP nữa: server chỉ lắng nghe 127.0.0.1 nên mọi request
    đều từ chính máy này.
    """
    from db import get_account_by_id
    row = get_account_by_id(acc_id)
    if not row:
        return jsonify({"ok": False, "error": "Không tìm thấy tài khoản"}), 404
    return jsonify({"ok": True, "data": {f: row.get(f, "") for f in SECRET_ACCOUNT_FIELDS}})


# Cột và nhãn khớp đúng bảng trên giao diện (ACC_FIELDS trong static/js/app.js).
EXPORT_ACCOUNT_COLUMNS = [
    ("ten_acc",        "Tên acc"),
    ("loai_dang",      "Loại đăng"),
    ("thoi_gian_nghi", "Nghỉ (p)"),
    ("link_profile",   "Link profile"),
    ("email_sdt",      "Email/SDT"),
    ("password",       "Password"),
    ("ten_page",       "Tên Page"),
    ("c_user",         "c_user"),
    ("xs",             "xs"),
    ("refresh",        "Refresh"),
    ("trang_thai",     "Trạng thái"),
    ("nuoi_nick",      "Nuôi"),
    ("nuoi_interval",  "Chu kỳ (p)"),
    ("email_khoiphuc", "Email KP"),
    ("pass_khoiphuc",  "Pass KP"),
    ("twofa",          "2FA"),
    ("ghi_chu",        "Ghi chú"),
]

# Excel tự nhận diện kiểu dữ liệu: c_user (15 chữ số) sẽ thành 1.00047E+14 và
# mất số gốc, thời gian nghỉ thành số nguyên. Ép các cột này về dạng Text.
EXPORT_TEXT_COLUMNS = {"password", "c_user", "xs", "twofa",
                       "thoi_gian_nghi", "nuoi_interval"}


def _export_dir() -> Path:
    """Thư mục lưu file xuất — Downloads của người dùng, không có thì cạnh app."""
    d = Path.home() / "Downloads"
    return d if d.is_dir() else BASE_DIR


def _reveal_in_explorer(path: Path):
    """Mở Explorer và chọn sẵn file vừa lưu (chỉ Windows, lỗi thì bỏ qua)."""
    if sys.platform != "win32":
        return
    try:
        # explorer trả exit code 1 kể cả khi thành công → không dùng check=True
        subprocess.Popen(f'explorer /select,"{path}"')
    except Exception as e:
        logger.warning(f"Không mở được Explorer: {e}")


@app.route("/api/accounts/export-excel")
def api_accounts_export_excel():
    """Xuất toàn bộ bảng Tài khoản ra .xlsx, lưu vào Downloads.

    CẢNH BÁO: cố ý KHÔNG che credential — file chứa mật khẩu, cookie xs và mã
    2FA của mọi tài khoản. Giữ file này cẩn thận.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Tài khoản"
    ws.append([label for _, label in EXPORT_ACCOUNT_COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    widths = [len(label) for _, label in EXPORT_ACCOUNT_COLUMNS]
    for row in get_accounts():
        values = []
        for key, _ in EXPORT_ACCOUNT_COLUMNS:
            v = row.get(key, "")
            if key == "nuoi_nick":
                v = "x" if v else ""
            values.append("" if v is None else str(v))
        ws.append(values)
        for i, v in enumerate(values):
            widths[i] = max(widths[i], len(v))

    for i, (key, _) in enumerate(EXPORT_ACCOUNT_COLUMNS, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        # +2 cho khoảng thở, trần 50 để link_profile/ghi_chu không kéo dài cả màn
        ws.column_dimensions[letter].width = min(widths[i - 1] + 2, 50)
        if key in EXPORT_TEXT_COLUMNS:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).number_format = "@"

    ws.freeze_panes = "A2"

    # Ghi thẳng xuống đĩa thay vì trả blob cho trình duyệt tải: app chạy trong
    # cửa sổ pywebview/WebView2, ở đó thẻ <a download> bị bỏ qua âm thầm nên
    # người dùng thấy báo thành công mà không có file nào.
    name = f"tai_khoan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = _export_dir() / name
    wb.save(str(path))

    _reveal_in_explorer(path)
    return jsonify({"ok": True, "path": str(path)})


@app.route("/api/accounts/import-excel", methods=["POST"])
def api_accounts_import_excel():
    """Nhập tài khoản từ file .xlsx (khớp định dạng file Xuất Excel).

    Chế độ "thêm & bỏ trùng": acc nào đã có (cùng Tên acc + Tên Page) thì bỏ qua,
    chỉ thêm acc mới. Cột nhận diện theo header ở dòng 1, không phụ thuộc thứ tự.
    """
    from openpyxl import load_workbook

    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "error": "Không có file"})

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return jsonify({"ok": False, "error": "File không phải Excel (.xlsx) hợp lệ"})
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return jsonify({"ok": False, "error": "File rỗng"})

    label_to_key = {label.strip().lower(): key for key, label in EXPORT_ACCOUNT_COLUMNS}
    col_key = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        key = label_to_key.get(str(h).strip().lower())
        if key:
            col_key[idx] = key

    if "ten_acc" not in col_key.values():
        return jsonify({"ok": False, "error": "File thiếu cột Tên acc"})

    records = []
    for row in rows_iter:
        rec = {}
        for idx, key in col_key.items():
            v = row[idx] if idx < len(row) else None
            v = "" if v is None else str(v).strip()
            # Cột "Nuôi" xuất ra là "x"/"" — đổi ngược về 1/0 cho DB.
            if key == "nuoi_nick":
                v = 1 if v.lower() == "x" else 0
            rec[key] = v
        if rec.get("ten_acc"):
            records.append(rec)

    try:
        added, skipped = import_accounts(records)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

    return jsonify({"ok": True, "added": added, "skipped": skipped})


@app.route("/api/accounts/save", methods=["POST"])
def api_accounts_save():
    data = request.json or {}
    # Form gửi lại dấu che nếu người dùng không sửa trường đó — bỏ qua để
    # không ghi đè credential thật bằng chuỗi "••••••".
    for f in SECRET_ACCOUNT_FIELDS:
        if data.get(f) == SECRET_MASK:
            data.pop(f)
    try:
        acc_id = upsert_account(data)
        return jsonify({"ok": True, "id": acc_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/accounts/<int:acc_id>/field", methods=["POST"])
def api_accounts_field(acc_id):
    body = request.json or {}
    if body.get("field") in SECRET_ACCOUNT_FIELDS and body.get("value") == SECRET_MASK:
        return jsonify({"ok": True, "skipped": True})   # không đổi gì
    try:
        update_account_field(acc_id, body["field"],
                             ep_kieu_so(body["field"], body["value"]))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── Cảnh báo sức khoẻ acc ───────────────────────────────────────────────
# Scheduler là tiến trình riêng nên không đẩy thẳng toast lên web được; nó ghi
# vào cột accounts.canh_bao_moi, giao diện hỏi ở đây rồi báo đã xem.
@app.route("/api/canh-bao")
def api_canh_bao():
    try:
        return jsonify({"ok": True, "data": db.lay_canh_bao()})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "data": []})


@app.route("/api/canh-bao/xong", methods=["POST"])
def api_canh_bao_xong():
    try:
        db.xoa_canh_bao()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/accounts/refresh-now", methods=["POST"])
def api_accounts_refresh_now():
    """
    Làm mới cookie NGAY cho các acc đang để cột Refresh = Yes.

    Vòng lặp scheduler cũng làm việc này, nhưng 10 phút mới quét một lần và chỉ
    chạy bên trong runner đang bật — dừng hết runner thì cột Refresh nằm ở Yes
    vĩnh viễn. Nút này gỡ cả hai ràng buộc.
    """
    try:
        from cookie_exporter import refresh_pending_accounts
        kq = refresh_pending_accounts()
        return jsonify({"ok": True, **kq})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/accounts/<int:acc_id>", methods=["DELETE"])
def api_accounts_delete(acc_id):
    delete_account(acc_id)
    return jsonify({"ok": True})


@app.route("/api/accounts/reorder", methods=["POST"])
def api_accounts_reorder():
    ordered_ids = request.json or []
    try:
        reorder_accounts([int(i) for i in ordered_ids])
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/pages/reorder", methods=["POST"])
def api_pages_reorder():
    ordered_ids = request.json or []
    try:
        reorder_pages([int(i) for i in ordered_ids])
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/accounts/<int:ref_id>/insert-row", methods=["POST"])
def api_accounts_insert_row(ref_id):
    position = (request.json or {}).get("position", "below")
    try:
        new_id = insert_account_at(ref_id, position)
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ═══════════════════════════════════════════════════════════════
# API — Pages
# ═══════════════════════════════════════════════════════════════

@app.route("/api/pages")
def api_pages():
    return jsonify({"ok": True, "data": get_pages()})


@app.route("/api/pages/save", methods=["POST"])
def api_pages_save():
    data = request.json or {}
    try:
        pid = upsert_page(data)
        return jsonify({"ok": True, "id": pid})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/pages/<int:page_id>/field", methods=["POST"])
def api_pages_field(page_id):
    body = request.json or {}
    field = body.get("field","")
    value = body.get("value","")
    safe = {"ten_page","acc_quan_ly","page_uid","link_page","loai_page","bai_dang_toi_da","ghi_chu"}
    if field not in safe:
        return jsonify({"ok": False, "error": f"Field không hợp lệ: {field}"})
    value = ep_kieu_so(field, value)
    try:
        from db import _conn
        with _conn() as con:
            con.execute(f"UPDATE pages SET {field}=? WHERE id=?", (value, page_id))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/pages/<int:page_id>", methods=["DELETE"])
def api_pages_delete(page_id):
    delete_page(page_id)
    return jsonify({"ok": True})


# Cột xuất/nhập Excel cho Page. Thứ tự này cũng là thứ tự cột trong file.
# Cột đầu là key DB, cột sau là header người đọc.
EXPORT_PAGE_COLUMNS = [
    ("ten_page",        "Tên Page"),
    ("acc_quan_ly",     "Acc quản lý"),
    ("page_uid",        "Page UID"),
    ("loai_page",       "Loại đăng"),
    ("bai_dang_toi_da", "Bài tối đa"),
    ("link_page",       "Link Page"),
    ("ghi_chu",         "Ghi chú"),
]


@app.route("/api/pages/export-excel")
def api_pages_export_excel():
    """Xuất danh sách Page ra .xlsx, lưu vào Downloads."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows = get_pages()

    wb = Workbook()
    ws = wb.active
    ws.title = "Page"
    ws.append([label for _, label in EXPORT_PAGE_COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    widths = [len(label) for _, label in EXPORT_PAGE_COLUMNS]
    for row in rows:
        values = []
        for key, _ in EXPORT_PAGE_COLUMNS:
            v = row.get(key, "")
            values.append("" if v is None else str(v))
        ws.append(values)
        for i, v in enumerate(values):
            widths[i] = max(widths[i], len(v))

    for i, (key, _) in enumerate(EXPORT_PAGE_COLUMNS, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = min(widths[i - 1] + 2, 50)
        # page_uid để dạng text tránh Excel biến số dài thành 1.23E+15.
        if key == "page_uid":
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).number_format = "@"

    ws.freeze_panes = "A2"

    name = f"page_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = _export_dir() / name
    wb.save(str(path))

    _reveal_in_explorer(path)
    return jsonify({"ok": True, "path": str(path), "count": len(rows)})


@app.route("/api/pages/import-excel", methods=["POST"])
def api_pages_import_excel():
    """Nhập Page từ file .xlsx do đồng nghiệp gửi.

    Chế độ "thêm & bỏ trùng": Page nào đã có thì bỏ qua, chỉ thêm Page mới.
    Trùng so theo Page UID; Page thiếu UID thì so thêm theo Tên Page.
    Cột nhận diện theo header ở dòng 1, không phụ thuộc thứ tự cột.
    """
    from openpyxl import load_workbook

    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "error": "Không có file"})

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return jsonify({"ok": False, "error": "File không phải Excel (.xlsx) hợp lệ"})
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return jsonify({"ok": False, "error": "File rỗng"})

    label_to_key = {label.strip().lower(): key for key, label in EXPORT_PAGE_COLUMNS}
    col_key = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        key = label_to_key.get(str(h).strip().lower())
        if key:
            col_key[idx] = key

    if "ten_page" not in col_key.values():
        return jsonify({"ok": False, "error": "File thiếu cột Tên Page"})

    records = []
    for row in rows_iter:
        rec = {}
        for idx, key in col_key.items():
            v = row[idx] if idx < len(row) else None
            rec[key] = "" if v is None else str(v).strip()
        if rec.get("ten_page"):
            records.append(rec)

    try:
        added, skipped = import_pages(records)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

    return jsonify({"ok": True, "added": added, "skipped": skipped})


# ═══════════════════════════════════════════════════════════════
# API — Content
# ═══════════════════════════════════════════════════════════════

@app.route("/api/content/<loai>")
def api_content(loai):
    rows = get_content(loai)
    return jsonify({"ok": True, "data": rows})


@app.route("/api/content/save", methods=["POST"])
def api_content_save():
    data = request.json or {}
    try:
        # Ảnh bị bỏ khỏi content khi sửa thì xóa luôn file, khỏi tồn đọng.
        from db import get_content_image_urls
        from storage import xoa_anh_khong_dung
        cu = get_content_image_urls(data["id"]) if data.get("id") else set()

        cid = upsert_content(data)

        bo_di = cu - get_content_image_urls(cid)
        n = xoa_anh_khong_dung(bo_di) if bo_di else 0
        return jsonify({"ok": True, "id": cid, "anh_da_xoa": n})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/content/<int:content_id>/field", methods=["POST"])
def api_content_field(content_id):
    body  = request.json or {}
    field = body.get("field", "")
    value = body.get("value", "")
    safe  = {"ma_content","noi_dung","link_anh","su_dung","ghi_chu"}
    if field not in safe:
        return jsonify({"ok": False, "error": f"Field không hợp lệ: {field}"})
    try:
        from db import _conn
        with _conn() as con:
            con.execute(f"UPDATE content SET {field}=? WHERE id=?", (value, content_id))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/content/<int:content_id>", methods=["DELETE"])
def api_content_delete(content_id):
    # Lấy danh sách ảnh TRƯỚC khi xóa dòng, rồi dọn file — trước đây chỉ xóa
    # dòng DB nên ảnh nằm lại trên đĩa mãi mãi.
    from db import get_content_image_urls
    from storage import xoa_anh_khong_dung
    anh = get_content_image_urls(content_id)
    delete_content(content_id)
    n = xoa_anh_khong_dung(anh) if anh else 0
    return jsonify({"ok": True, "anh_da_xoa": n})


@app.route("/api/content/reorder", methods=["POST"])
def api_content_reorder():
    ordered_ids = request.json or []
    try:
        reorder_content([int(i) for i in ordered_ids])
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/content/quet-anh-mo-coi", methods=["POST"])
def api_quet_anh_mo_coi():
    """Quét (và tuỳ chọn xóa) ảnh không content nào dùng tới."""
    from storage import quet_anh_mo_coi
    xoa = bool((request.json or {}).get("xoa"))
    try:
        return jsonify({"ok": True, **quet_anh_mo_coi(xoa=xoa)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/content/upload-image", methods=["POST"])
def api_content_upload_image():
    """Upload ảnh lên local storage."""
    from storage import save_image
    loai = request.form.get("loai", "uploads")
    file = request.files.get("image")
    if not file:
        return jsonify({"ok": False, "error": "Không có file"})
    try:
        url = save_image(file.read(), file.filename, loai)
        return jsonify({"ok": True, "url": url})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# Danh mục (tab) của Content — CỐ ĐỊNH 3 loại, khớp với phần Lịch (homestay/thue/ban).
# Không cho thêm/xóa/sửa tab để giữ logic đơn giản, đồng bộ với scheduler.
CONTENT_CATEGORIES = [
    {"key": "homestay", "title": "Homestay", "icon": "🏠"},
    {"key": "thue",     "title": "Thuê",     "icon": "🏡"},
    {"key": "ban",      "title": "Bán",      "icon": "💰"},
]


@app.route("/api/content-categories")
def api_content_categories():
    return jsonify({"ok": True, "data": CONTENT_CATEGORIES})


# ═══════════════════════════════════════════════════════════════
# API — UID Groups
# ═══════════════════════════════════════════════════════════════

@app.route("/api/uid-groups")
def api_uid_groups():
    return jsonify({"ok": True, "data": get_all_uid_groups()})


@app.route("/api/uid-groups/save", methods=["POST"])
def api_uid_groups_save():
    data = request.json or {}
    try:
        gid = upsert_uid_group(data)
        return jsonify({"ok": True, "id": gid})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/uid-groups/reorder", methods=["POST"])
def api_uid_groups_reorder():
    ordered_ids = request.json or []
    try:
        reorder_uid_groups([int(i) for i in ordered_ids])
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/uid-groups/<int:gid>", methods=["DELETE"])
def api_uid_groups_delete(gid):
    delete_uid_group(gid)
    return jsonify({"ok": True})


# Cột xuất/nhập Excel cho UID nhóm. Thứ tự này cũng là thứ tự cột trong file.
# Cột đầu là header người đọc, key là tên cột DB tương ứng.
EXPORT_UID_COLUMNS = [
    ("uid",        "UID"),
    ("ten_nhom",   "Tên nhóm"),
    ("link_url",   "Link"),
    ("thanh_vien", "Thành viên"),
    ("ghi_chu",    "Ghi chú"),
]


@app.route("/api/uid-groups/export-excel")
def api_uid_groups_export_excel():
    """Xuất danh sách UID nhóm ra .xlsx, lưu vào Downloads.

    Chỉ xuất nhóm từ sheet "UID Nhóm" (ma_nhom trống) — đúng những gì hiển thị
    trên tab UID Nhóm. Các mã TIME1-7 dùng nội bộ nên không đưa vào file chia sẻ.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows = [g for g in get_all_uid_groups() if not g.get("ma_nhom")]

    wb = Workbook()
    ws = wb.active
    ws.title = "UID Nhóm"
    ws.append([label for _, label in EXPORT_UID_COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    widths = [len(label) for _, label in EXPORT_UID_COLUMNS]
    for row in rows:
        values = []
        for key, _ in EXPORT_UID_COLUMNS:
            v = row.get(key, "")
            values.append("" if v is None else str(v))
        ws.append(values)
        for i, v in enumerate(values):
            widths[i] = max(widths[i], len(v))

    for i, (key, _) in enumerate(EXPORT_UID_COLUMNS, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = min(widths[i - 1] + 2, 50)
        # UID để dạng text tránh Excel biến số dài thành 1.23E+15.
        if key == "uid":
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).number_format = "@"

    ws.freeze_panes = "A2"

    name = f"uid_nhom_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = _export_dir() / name
    wb.save(str(path))

    _reveal_in_explorer(path)
    return jsonify({"ok": True, "path": str(path), "count": len(rows)})


@app.route("/api/uid-groups/import-excel", methods=["POST"])
def api_uid_groups_import_excel():
    """Nhập UID nhóm từ file .xlsx do đồng nghiệp gửi.

    Chế độ "thêm & bỏ trùng": UID nào đã có (so theo cột uid, sheet UID Nhóm)
    thì bỏ qua, chỉ thêm UID mới. Không đụng tới dữ liệu cũ.
    Cột nhận diện theo header ở dòng 1, không phụ thuộc thứ tự cột.
    """
    from openpyxl import load_workbook

    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "error": "Không có file"})

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return jsonify({"ok": False, "error": "File không phải Excel (.xlsx) hợp lệ"})
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return jsonify({"ok": False, "error": "File rỗng"})

    # Map tên header (chuẩn hoá thường, bỏ khoảng trắng) -> chỉ số cột.
    label_to_key = {label.strip().lower(): key for key, label in EXPORT_UID_COLUMNS}
    col_key = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        key = label_to_key.get(str(h).strip().lower())
        if key:
            col_key[idx] = key

    if "uid" not in col_key.values():
        return jsonify({"ok": False, "error": "File thiếu cột UID"})

    records = []
    for row in rows_iter:
        rec = {}
        for idx, key in col_key.items():
            v = row[idx] if idx < len(row) else None
            rec[key] = "" if v is None else str(v).strip()
        if rec.get("uid"):
            records.append(rec)

    try:
        added, skipped = import_uid_groups(records)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

    return jsonify({"ok": True, "added": added, "skipped": skipped})


# ═══════════════════════════════════════════════════════════════
# API — Schedules
# ═══════════════════════════════════════════════════════════════

@app.route("/api/schedule/<loai>")
def api_schedule(loai):
    rows = get_schedules(loai)
    return jsonify({"ok": True, "data": rows})


@app.route("/api/schedule/<loai>/reset", methods=["POST"])
def api_schedule_reset(loai):
    count = bulk_set_schedule_status(loai, "✅", "Chờ")
    count += bulk_set_schedule_status(loai, "❌", "Chờ")
    count += bulk_set_schedule_status(loai, "X", "Chờ")
    return jsonify({"ok": True, "updated": count})


@app.route("/api/schedule/<loai>/stop", methods=["POST"])
def api_schedule_stop(loai):
    count = bulk_set_schedule_status(loai, "Chờ", "X")
    return jsonify({"ok": True, "updated": count})


@app.route("/api/schedule/<loai>/cell", methods=["POST"])
def api_schedule_cell(loai):
    body = request.json or {}
    try:
        update_schedule_field(body["id"], body["field"], body["value"])
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/schedule/<loai>/gen", methods=["POST"])
def api_schedule_gen(loai):
    """Gen lịch đăng chéo (Homestay/Thuê/Bán) và lưu vào SQLite."""
    body         = request.json or {}
    start_str    = body.get("start", "05:00")
    end_str      = body.get("end",   "03:00")
    keyword_pool = body.get("keyword_pool", [])
    acc_settings = body.get("acc_settings", [])

    def parse_hhmm(s, dh, dm):
        try:
            h, m = map(int, s.split(":"))
            return h * 60 + m
        except Exception:
            return dh * 60 + dm

    def min_to_time(m):
        return f"{(m // 60) % 24:02d}:{m % 60:02d}"

    START_MIN = parse_hhmm(start_str, 5, 0)
    end_raw   = parse_hhmm(end_str, 3, 0)
    END_MIN   = end_raw if end_raw > START_MIN else end_raw + 24 * 60

    if not acc_settings:
        return jsonify({"ok": False, "error": "Không có acc nào"})

    # "Nhóm đầu" (UID/link mở composer) là bắt buộc — fail sớm thay vì lỗi lúc chạy.
    thieu = [a["ten"] for a in acc_settings if not (a.get("first_group_url", "") or "").strip()]
    if thieu:
        return jsonify({"ok": False,
                        "error": "Thiếu 'Nhóm đầu' cho acc: " + ", ".join(thieu)})

    n_kw   = len(keyword_pool) if keyword_pool else 1
    n_accs = len(acc_settings)

    contents = acc_settings[0].get("contents", []) if acc_settings else []
    if not contents:
        # Lấy từ DB
        from db import get_content
        contents = [r["ma_content"] for r in get_content(loai, su_dung="Có")]
    if not contents:
        return jsonify({"ok": False, "error": f"Không có content '{loai}' nào (Sử dụng=Có)"})

    n_contents     = len(contents)
    content_cursor = {a["ten"]: a.get("c_offset", 0) for a in acc_settings}
    keyword_cursor = {a["ten"]: (i * n_kw // n_accs) % n_kw for i, a in enumerate(acc_settings)}
    theo_ten       = {a["ten"]: a for a in acc_settings}

    # Phân bổ thời điểm — phiên comment tính ngang phiên đăng bài, cùng một
    # vòng xoay, cùng góp vào tổng lực. Xem xep_lich.py để hiểu vì sao phải ép
    # giãn cách tối thiểu.
    from xep_lich import phan_bo_lich
    moc = phan_bo_lich(acc_settings, START_MIN, END_MIN)

    schedule = []
    for t_min, ten in moc:
        acc = theo_ten[ten]

        cur_c = content_cursor[ten]
        content_cursor[ten] = (cur_c + 1) % n_contents

        mode_acc = acc.get("mode", "Hybrid")
        cur_kw   = keyword_cursor[ten]
        if keyword_pool and mode_acc in ("Via", "Hybrid"):
            tu_khoa = keyword_pool[cur_kw % n_kw]
            keyword_cursor[ten] = (cur_kw + 1) % n_kw
        else:
            tu_khoa = ""

        # ma_nhom = URL/UID nhóm đầu để mở composer (đã validate là luôn có).
        ma_nhom_val = acc.get("first_group_url", "").strip()

        schedule.append({
            "loai":       loai,
            "stt":        len(schedule) + 1,
            "ma_content": contents[cur_c % n_contents],
            "ten_acc":    ten,
            "ten_page":   acc["page"],
            "gio_dang":   min_to_time(t_min),
            "ma_nhom":    ma_nhom_val,
            "tu_khoa":    tu_khoa,
            "mode":       mode_acc,
            "trang_thai": "Chờ",
        })

    if not schedule:
        return jsonify({"ok": False, "error": "Không tạo được lịch"})

    # ── Nuôi nick: chuyển một số slot của acc bật nuôi thành phiên nuôi ──
    # Nick càng non → chuyển càng nhiều (đăng ít, nuôi nhiều). Đọc thẳng từ DB
    # theo tên acc nên không cần đổi luồng gen ở frontend.
    #
    # `da_dat` gom mốc giờ của MỌI phiên đã chuyển và được truyền tiếp sang lượt
    # comment bên dưới, để phiên nuôi và phiên comment không rơi sát nhau.
    n_warm, n_cmt = 0, 0
    da_dat = []
    try:
        from nuoi_nick import plan_warming_conversion
        acc_names = {r["ten_acc"] for r in schedule}
        accs_db   = [a for a in get_accounts() if a["ten_acc"] in acc_names]

        chi_cmt = {a["ten_acc"] for a in accs_db
                   if la_loai_comment(a.get("loai_dang"))}

        # ── Thứ tự ba bước dưới đây là có chủ đích ──
        # Cả hai hàm chuyển slot chỉ đụng slot đang là 'dang_bai'. Vì vậy nuôi
        # nick phải chạy TRƯỚC, rồi mới quét nốt slot còn lại của acc C_* thành
        # comment. Làm ngược lại thì acc C_* bị khoá hết slot thành 'comment' và
        # KHÔNG BAO GIỜ được nuôi — trong khi acc chỉ comment cũng cần nuôi y
        # như acc đăng bài: hy sinh một phiên comment để đi nuôi.

        # 1. Nuôi nick — áp cho MỌI acc tick Nuôi, kể cả acc C_* chỉ comment.
        warm_accs = {a["ten_acc"]: a.get("nuoi_interval")
                     for a in accs_db if int(a.get("nuoi_nick", 0) or 0) == 1}
        if warm_accs:
            n_warm = plan_warming_conversion(schedule, warm_accs, da_dat=da_dat)

        # 2. Acc "X_*" — vừa đăng vừa comment theo TỈ LỆ (mặc định 75/25).
        #    Chỉ 3 loại có danh sách bài + thư viện câu; lịch Page thì slot
        #    comment tới giờ chỉ mở trình duyệt rồi bỏ qua — mất slot đăng mà
        #    chẳng làm được gì.
        if loai in LOAI_LICH_MAP:
            from xep_lich import chuyen_slot_theo_ti_le
            hon_hop = {a["ten_acc"] for a in accs_db
                       if la_loai_hon_hop(a.get("loai_dang"))}
            if hon_hop:
                ti_le = int(get_setting("comment_ti_le",
                                        str(TI_LE_COMMENT_MAC_DINH))
                            or TI_LE_COMMENT_MAC_DINH)
                n_cmt += chuyen_slot_theo_ti_le(schedule, hon_hop, ti_le)

        # 3. Acc C_*: mọi slot CÒN LẠI thành comment (slot đã bị nuôi chiếm ở
        #    bước 1 thì giữ nguyên là phiên nuôi).
        for r in schedule:
            if r["ten_acc"] in chi_cmt and \
                    (r.get("hoat_dong") or "dang_bai") == "dang_bai":
                r["hoat_dong"] = "comment"
                n_cmt += 1
    except Exception as e:
        logger.warning(f"Chuyển slot nuôi/comment: bỏ qua ({e})")

    replace_schedules(loai, schedule)

    # Lưu thiết lập gần nhất để lần gen sau prefill (thay cho hardcode mặc định).
    set_setting(f"gen_prefs_{loai}", json.dumps({
        "start":       start_str,
        "end":         end_str,
        "first_group": (acc_settings[0].get("first_group_url", "") or "").strip(),
        "keyword":     ",".join(keyword_pool),
    }, ensure_ascii=False))

    return jsonify({"ok": True, "total": len(schedule),
                    "nuoi": n_warm,
                    "comment": n_cmt,
                    "from": schedule[0]["gio_dang"],
                    "to":   schedule[-1]["gio_dang"]})


@app.route("/api/nuoi/msg-mau")
def api_nuoi_msg_mau():
    """Thư viện câu nhắn mẫu (nuoi_msg_mau.txt) để nạp vào ô Thư viện câu."""
    f = BASE_DIR / "nuoi_msg_mau.txt"
    if not f.exists():
        return jsonify({"ok": False, "error": "Không tìm thấy nuoi_msg_mau.txt"})
    lines = [ln.strip() for ln in f.read_text(encoding="utf-8").splitlines()]
    lines = [ln for ln in lines if ln and not ln.startswith("#")]
    return jsonify({"ok": True, "total": len(lines), "text": "\n".join(lines)})


@app.route("/api/comment/cau-mau")
def api_comment_cau_mau():
    """
    Thư viện câu comment mẫu (comment_mau.txt), tách theo loại đăng.

    Nạp tay từng câu trên mỗi máy vệ tinh vừa tốn thời gian vừa dễ ra kết quả
    khác nhau giữa các máy — mà ba bộ câu giống nhau chính là kiểu trùng lặp dễ
    bị quét nhất. Để câu mẫu đi theo repo thì mọi máy nạp ra cùng một bộ.
    """
    f = BASE_DIR / "comment_mau.txt"
    if not f.exists():
        return jsonify({"ok": False, "error": "Không tìm thấy comment_mau.txt"})
    data, loai = {}, None
    for ln in f.read_text(encoding="utf-8").splitlines():
        ln = ln.strip()
        if not ln or ln.startswith("#"):
            continue
        if ln.startswith("[") and ln.endswith("]"):
            loai = ln[1:-1].strip().lower()
            data.setdefault(loai, [])
            continue
        if loai:
            data[loai].append(ln)
    return jsonify({"ok": True,
                    "data":  {k: "\n".join(v) for k, v in data.items()},
                    "total": {k: len(v) for k, v in data.items()}})


@app.route("/api/schedule/nuoi/gen", methods=["POST"])
def api_schedule_nuoi_gen():
    """
    Gen lịch cho acc CHỈ NUÔI — acc có tick 'Nuôi' nhưng cột 'Loại đăng' để trống
    (không được phân công đăng bài). Mỗi acc vào một phiên nuôi mỗi `nuoi_interval`
    phút; các acc lệch pha nhau để không cùng mở trình duyệt một lúc.
    """
    from nuoi_nick import build_warming_schedule, normalize_interval
    body  = request.json or {}
    start = body.get("start", "07:00")
    end   = body.get("end",   "23:00")

    accs = [{"ten": a["ten_acc"], "interval": normalize_interval(a.get("nuoi_interval"))}
            for a in get_accounts(trang_thai="Active")
            if int(a.get("nuoi_nick", 0) or 0) == 1
            and not (a.get("loai_dang") or "").strip()]

    if not accs:
        return jsonify({"ok": False,
                        "error": "Không có acc nào 'chỉ nuôi' "
                                 "(cần: tick Nuôi + để TRỐNG cột Loại đăng + Trạng thái Active)"})

    rows = build_warming_schedule(accs, start, end)
    if not rows:
        return jsonify({"ok": False, "error": "Không tạo được lịch nuôi"})

    replace_schedules("nuoi", rows)
    set_setting("gen_prefs_nuoi", json.dumps({"start": start, "end": end}, ensure_ascii=False))
    return jsonify({"ok": True, "total": len(rows), "accs": len(accs),
                    "from": rows[0]["gio_dang"], "to": rows[-1]["gio_dang"]})


@app.route("/api/schedule/page/gen", methods=["POST"])
def api_schedule_page_gen():
    """Gen lịch Đăng bài Page từ bảng pages + content local."""
    import random as _random
    body    = request.json or {}
    acc     = body.get("acc", "")
    start_h = int(body.get("start_hour", 7))
    end_h   = int(body.get("end_hour", 23))

    CONTENT_LOAI_MAP = {"Homestay": "homestay", "Thuê": "thue", "Bán": "ban"}

    # Ép int khi ĐỌC, không so sánh thẳng: dữ liệu cũ (hoặc nhập từ Excel) có
    # thể còn lưu dạng chuỗi, '' > 0 sẽ ném TypeError -> HTTP 500.
    def _so_bai(p):
        return ep_kieu_so("bai_dang_toi_da", p.get("bai_dang_toi_da", 0))

    pages = [p for p in get_pages() if _so_bai(p) > 0]
    if not pages:
        return jsonify({"ok": False, "error": "Không có Page nào có 'Bài đăng tối đa' > 0"})

    page_items = []
    for p in pages:
        ct_loai = CONTENT_LOAI_MAP.get(p.get("loai_page", ""), "")
        if not ct_loai:
            continue
        codes = [r["ma_content"] for r in get_content(ct_loai, su_dung="Có")]
        if not codes:
            continue
        n      = min(_so_bai(p), len(codes))
        picked = _random.sample(codes, n)
        page_items.append({"ten_page": p["ten_page"], "picked": picked})

    if not page_items:
        return jsonify({"ok": False, "error": "Không có content nào"})

    def interleave(lists):
        res = []; iters = [list(l) for l in lists]
        while any(iters):
            for l in iters:
                if l: res.append(l.pop(0))
        return res

    mixed    = interleave([[(p["ten_page"], c) for c in p["picked"]] for p in page_items])
    total    = len(mixed)
    total_min = (end_h - start_h) * 60
    interval  = total_min / (total - 1) if total > 1 else 0

    schedule = []
    for i, (ten_page, ma_content) in enumerate(mixed):
        t = start_h * 60 + round(i * interval)
        schedule.append({
            "loai":       "page",
            "stt":        i + 1,
            "ma_content": ma_content,
            "ten_acc":    acc,
            "ten_page":   ten_page,
            "gio_dang":   f"{(t//60)%24:02d}:{t%60:02d}",
            "ma_nhom":    "",
            "tu_khoa":    "",
            "mode":       "PAGE",
            "trang_thai": "Chờ",
        })

    replace_schedules("page", schedule)

    # Lưu thiết lập gen Page gần nhất để lần sau prefill.
    set_setting("gen_prefs_page", json.dumps({
        "acc":        acc,
        "start_hour": start_h,
        "end_hour":   end_h,
    }, ensure_ascii=False))

    return jsonify({"ok": True, "total": len(schedule),
                    "from": schedule[0]["gio_dang"],
                    "to":   schedule[-1]["gio_dang"]})


def _hhmm_to_min(s: str, mac_dinh: int) -> int:
    """'05:00' → 300. Chuỗi hỏng thì lấy mặc định."""
    try:
        h, m = map(int, str(s).split(":"))
        return h * 60 + m
    except Exception:
        return mac_dinh


@app.route("/api/schedule/<loai>/gen-data")
def api_schedule_gen_data(loai):
    """Lấy data cho form gen lịch: acc active + content pool."""
    # Gồm cả acc đăng bài (Homestay/Thuê/Bán) lẫn acc chỉ comment (C_Home/...).
    # Acc chỉ comment vẫn cần slot trong lịch — slot của họ sẽ được đánh dấu
    # hoat_dong='comment' ở bước gen bên dưới.
    accs_db = (accounts_theo_lich(loai) if loai in LOAI_LICH_MAP
               else get_accounts(loai=loai.capitalize(), trang_thai="Active"))

    accs = []
    for a in accs_db:
        try:
            nghi = int(a.get("thoi_gian_nghi", 30))
        except Exception:
            nghi = 30
        accs.append({
            "ten":      a["ten_acc"],
            "page":     a.get("ten_page", ""),
            "nghi":     nghi,
            "luc_dang": round(60 / nghi, 2),
            # Acc C_* chỉ đi comment — form gen ẩn ô Content/Mode cho họ, và
            # bảng lịch hiện badge 💬 thay vì mã content.
            "chi_comment": la_loai_comment(a.get("loai_dang")),
            "hon_hop":     la_loai_hon_hop(a.get("loai_dang")),
        })

    contents = [r["ma_content"] for r in get_content(loai, su_dung="Có")]

    # Số liệu phủ đều để người dùng thấy trước khi bấm Gen.
    from xep_lich import tong_luc as _tl, do_nen as _dn
    nhip = {"tong_luc": round(_tl(accs), 2), "do_nen": round(_dn(accs), 2),
            "luc_dang":   round(sum(a["luc_dang"] for a in accs if not a["chi_comment"]), 2),
            "luc_comment": round(sum(a["luc_dang"] for a in accs if a["chi_comment"]), 2)}

    # Thiết lập gen gần nhất đã lưu (link nhóm đầu, từ khóa, giờ) — None nếu chưa có.
    prefs = None
    raw = get_setting(f"gen_prefs_{loai}", "")
    if raw:
        try:
            prefs = json.loads(raw)
        except Exception:
            prefs = None

    return jsonify({"ok": True, "accs": accs, "contents": contents,
                    "prefs": prefs, "nhip": nhip})


# ═══════════════════════════════════════════════════════════════
# API — Logs
# ═══════════════════════════════════════════════════════════════

def _read_log(filename, n=150):
    f = Path(filename)
    if not f.exists():
        return "(Chưa có log)"
    try:
        lines = f.read_text(encoding="utf-8", errors="replace").splitlines()
        return "\n".join(lines[-n:])
    except Exception as e:
        return f"Lỗi: {e}"


@app.route("/api/logs/<loai>")
def api_logs(loai):
    n = int(request.args.get("n", 150))
    log_map = {
        "homestay": str(LOG_DIR / "autopost_homestay.log"),
        "thue":     str(LOG_DIR / "autopost_thue.log"),
        "ban":      str(LOG_DIR / "autopost_ban.log"),
        "page":     str(LOG_DIR / "autopost_page.log"),
        "nuoi":     str(LOG_DIR / "autopost_nuoi.log"),
    }
    fname = log_map.get(loai)
    if not fname:
        return jsonify({"ok": False, "error": "Loại không hợp lệ"})
    return jsonify({"ok": True, "text": _read_log(fname, n)})


# ═══════════════════════════════════════════════════════════════
# API — Bài viết để đi comment
# ═══════════════════════════════════════════════════════════════

@app.route("/api/comment-posts/<loai>")
def api_comment_posts(loai):
    """
    Kèm tên Page đã đăng bài. DB chỉ lưu `page_uid` (biết chắc lúc thu link),
    còn tên Page đổi được nên tra ngược lúc hiển thị chứ không lưu cứng.
    """
    ten_theo_uid = {str(p.get("page_uid") or ""): p.get("ten_page", "")
                    for p in get_pages() if p.get("page_uid")}
    rows = get_comment_posts(loai)
    for r in rows:
        uid = str(r.get("page") or "")
        r["ten_page"] = ten_theo_uid.get(uid, uid)
    return jsonify({"ok": True, "data": rows})


@app.route("/api/comment-posts/<loai>/add", methods=["POST"])
def api_comment_posts_add(loai):
    """Dán cả danh sách: mỗi dòng một URL. Bỏ dòng không phải link và URL trùng."""
    raw  = (request.json or {}).get("urls", "")
    urls = [u.strip() for u in raw.replace(",", "\n").splitlines()]
    urls = [u for u in urls if u.startswith("http")]
    if not urls:
        return jsonify({"ok": False, "error": "Không có link hợp lệ (phải bắt đầu bằng http)"})
    them = them_comment_posts(loai, urls)
    return jsonify({"ok": True, "them": them, "bo_trung": len(urls) - them})


@app.route("/api/comment-posts/<int:post_id>/field", methods=["POST"])
def api_comment_post_field(post_id):
    body = request.json or {}
    try:
        update_comment_post_field(post_id, body.get("field", ""), body.get("value", ""))
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)})
    return jsonify({"ok": True})


@app.route("/api/comment-posts/<int:post_id>", methods=["DELETE"])
def api_comment_post_delete(post_id):
    delete_comment_post(post_id)
    return jsonify({"ok": True})


@app.route("/api/comment-posts/<loai>/clear", methods=["POST"])
def api_comment_posts_clear(loai):
    return jsonify({"ok": True, "da_xoa": xoa_het_comment_posts(loai)})


# ═══════════════════════════════════════════════════════════════
# API — Settings
# ═══════════════════════════════════════════════════════════════

@app.route("/api/settings")
def api_settings():
    return jsonify({"ok": True, "data": get_all_settings()})


@app.route("/api/settings/save", methods=["POST"])
def api_settings_save():
    body = request.json or {}
    for k, v in body.items():
        set_setting(k, str(v))
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════
# Entry
# ═══════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════
# API — Tham gia nhóm
# ═══════════════════════════════════════════════════════════════

JOIN_LOG_FILE = str(LOG_DIR / "join_groups.log")

def _join_pid_file(sched_id: int) -> Path:
    return BASE_DIR / f".runner_join_{sched_id}.pid"

def _join_pid_alive(pid: int) -> bool:
    return _pid_alive(pid)

def _join_running_for(sched_id: int) -> bool:
    pf = _join_pid_file(sched_id)
    if not pf.exists():
        return False
    try:
        pid = int(pf.read_text().strip())
        if pid and _join_pid_alive(pid):
            return True
    except Exception:
        pass
    pf.unlink(missing_ok=True)
    return False

def _any_join_running() -> bool:
    return any(_join_running_for(int(pf.stem.split("_")[-1]))
               for pf in BASE_DIR.glob(".runner_join_*.pid"))

def _reset_stale_join():
    """Reset từng schedule bị stuck 'Đang chạy' khi process đã chết."""
    try:
        from db import _conn
        rows = _conn().execute(
            "SELECT id FROM join_schedules WHERE trang_thai='Đang chạy'"
        ).fetchall()
        for row in rows:
            sid = row[0]
            if not _join_running_for(sid):
                with _conn() as con:
                    con.execute("UPDATE join_schedules SET trang_thai='Chờ' WHERE id=?", (sid,))
    except Exception:
        pass


@app.route("/api/join/schedules")
def api_join_schedules():
    try:
        _reset_stale_join()
        from db import _conn
        rows = [dict(r) for r in _conn().execute(
            "SELECT * FROM join_schedules ORDER BY id DESC"
        ).fetchall()]
        # Gắn trạng thái running per-row
        for r in rows:
            r["is_running"] = _join_running_for(r["id"])
        return jsonify({"ok": True, "data": rows, "running": _any_join_running()})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/join/add", methods=["POST"])
def api_join_add():
    body     = request.json or {}
    ten_acc  = body.get("ten_acc","").strip()
    ten_page = body.get("ten_page","").strip()
    gio_chay = body.get("gio_chay","").strip()
    if not ten_acc or not ten_page:
        return jsonify({"ok": False, "error": "Thiếu Acc hoặc Page"})
    try:
        from db import _conn, get_page_by_name
        page_info = get_page_by_name(ten_page)
        page_uid  = page_info.get("page_uid","") if page_info else ""
        with _conn() as con:
            cur = con.execute(
                "INSERT INTO join_schedules (ten_acc,ten_page,page_uid,gio_chay,created_at) VALUES (?,?,?,?,datetime('now','localtime'))",
                (ten_acc, ten_page, page_uid, gio_chay)
            )
        return jsonify({"ok": True, "id": cur.lastrowid})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/join/gen-quick", methods=["POST"])
def api_join_gen_quick():
    """Tạo lịch nhanh: mỗi acc Active có page → 1 lịch tham gia nhóm."""
    body       = request.json or {}
    delay_new  = int(body.get("delay_new",  30))
    delay_skip = int(body.get("delay_skip",  5))
    # Lưu settings để dùng khi Run
    set_setting("join_delay_new",  str(delay_new))
    set_setting("join_delay_skip", str(delay_skip))

    try:
        accs    = get_accounts(trang_thai="Active")
        created = 0
        skipped = 0
        from db import _conn, get_page_by_name
        with _conn() as con:
            for acc in accs:
                ten_acc  = acc["ten_acc"]
                ten_page = (acc.get("ten_page") or "").strip()
                if not ten_page:
                    continue
                # Bỏ qua nếu đã có lịch cho cặp này
                existing = con.execute(
                    "SELECT id FROM join_schedules WHERE ten_acc=? AND ten_page=?",
                    (ten_acc, ten_page)
                ).fetchone()
                if existing:
                    skipped += 1
                    continue
                page_info = get_page_by_name(ten_page)
                page_uid  = page_info.get("page_uid", "") if page_info else ""
                con.execute(
                    "INSERT INTO join_schedules (ten_acc,ten_page,page_uid,gio_chay,trang_thai,created_at) "
                    "VALUES (?,?,?,?,?,datetime('now','localtime'))",
                    (ten_acc, ten_page, page_uid, "", "Chờ")
                )
                created += 1
        return jsonify({"ok": True, "created": created, "skipped": skipped})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/join/<int:sched_id>/run", methods=["POST"])
def api_join_run(sched_id):
    # Cho phép song song — chỉ chặn nếu chính schedule này đang chạy
    if _join_running_for(sched_id):
        return jsonify({"ok": False, "error": f"Lịch #{sched_id} đang chạy rồi"})
    try:
        from db import _conn
        row = _conn().execute("SELECT * FROM join_schedules WHERE id=?", (sched_id,)).fetchone()
        if not row: return jsonify({"ok": False, "error": "Không tìm thấy lịch"})
        row = dict(row)
        body       = request.json or {}
        headless   = body.get("headless", True)
        delay_new  = int(body.get("delay_new",  get_setting("join_delay_new",  "30") or "30"))
        delay_skip = int(body.get("delay_skip", get_setting("join_delay_skip",  "5") or "5"))
        flags      = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
        env        = {**os.environ,
                      "JOIN_SCHEDULE_ID":   str(sched_id),
                      "JOIN_ACC_NAME":      row["ten_acc"],
                      "JOIN_PAGE_UID":      row["page_uid"],
                      "HEADLESS":           "true" if headless else "false",
                      "JOIN_DELAY_NEW":     str(delay_new),
                      "JOIN_DELAY_SKIP":    str(delay_skip),
                      "SCHEDULER_LOG_FILE": JOIN_LOG_FILE}
        with _conn() as con:
            con.execute("UPDATE join_schedules SET trang_thai='Đang chạy', moi_join=0, da_join=0, loi=0, tong_nhom=0 WHERE id=?", (sched_id,))
        proc = subprocess.Popen([sys.executable, "-X", "utf8", "join_groups_worker.py"],
                                cwd=str(BASE_DIR), creationflags=flags, env=env)
        _join_pid_file(sched_id).write_text(str(proc.pid))
        return jsonify({"ok": True, "pid": proc.pid})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/join/<int:sched_id>/stop", methods=["POST"])
def api_join_stop(sched_id):
    pf = _join_pid_file(sched_id)
    killed = False
    if pf.exists():
        try:
            pid = int(pf.read_text().strip())
            subprocess.run(["taskkill", "/F", "/T", "/PID", str(pid)], capture_output=True)
            killed = True
        except Exception:
            pass
        pf.unlink(missing_ok=True)
    try:
        from db import _conn
        with _conn() as con:
            con.execute("UPDATE join_schedules SET trang_thai='Chờ' WHERE id=?", (sched_id,))
    except Exception:
        pass
    return jsonify({"ok": True, "killed": killed})


@app.route("/api/join/<int:sched_id>/status")
def api_join_status(sched_id):
    try:
        from db import _conn
        row = _conn().execute("SELECT * FROM join_schedules WHERE id=?", (sched_id,)).fetchone()
        if not row: return jsonify({"ok": False, "error": "Không tìm thấy"})
        return jsonify({"ok": True, "data": dict(row), "running": _join_running_for(sched_id)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/join/<int:sched_id>", methods=["DELETE"])
def api_join_delete(sched_id):
    try:
        from db import _conn
        with _conn() as con:
            con.execute("DELETE FROM join_schedules WHERE id=?", (sched_id,))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/logs/join")
def api_logs_join():
    n = int(request.args.get("n", 200))
    return jsonify({"ok": True, "text": _read_log(JOIN_LOG_FILE, n)})


def _wait_server_ready(port: int, timeout: float = 15.0):
    """Chờ Flask nhận kết nối trước khi mở cửa sổ (tránh trang lỗi lúc đầu)."""
    import socket
    end = time.time() + timeout
    while time.time() < end:
        try:
            with socket.create_connection(("127.0.0.1", port), timeout=0.5):
                return True
        except OSError:
            time.sleep(0.2)
    return False


def _serve():
    # 127.0.0.1 — CHỈ máy này truy cập được. Trước đây bind 0.0.0.0 để điều
    # khiển từ xa, kéo theo việc mọi máy trong mạng LAN cũng chạm được cổng 8080.
    app.run(host="127.0.0.1", port=PORT, debug=False, use_reloader=False, threaded=True)


if __name__ == "__main__":
    import threading
    _kill_all_runners()

    # --browser  : mở bằng trình duyệt mặc định (chế độ web cũ, để debug)
    # --no-browser: chỉ chạy server, không mở gì (server chạy ẩn)
    # (mặc định)  : mở CỬA SỔ APP riêng bằng pywebview — tách biệt với Chrome
    if "--browser" in sys.argv:
        import webbrowser
        webbrowser.open(f"http://localhost:{PORT}")
        _serve()
    elif "--no-browser" in sys.argv:
        _serve()
    else:
        try:
            import webview
        except ImportError:
            # Chưa cài pywebview → fallback mở trình duyệt như cũ
            import webbrowser
            webbrowser.open(f"http://localhost:{PORT}")
            _serve()
        else:
            # Icon riêng (chữ MNT xanh) — tách khỏi icon pythonw mặc định
            _icon = str(BASE_DIR / "static" / "mnt.ico")
            if sys.platform == "win32":
                try:
                    import ctypes
                    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("MNT.AutoPost")
                except Exception:
                    pass

            # Flask chạy nền, cửa sổ app giữ thread chính
            threading.Thread(target=_serve, daemon=True).start()
            _wait_server_ready(PORT)
            win = webview.create_window(
                "MNT AutoPost",
                f"http://localhost:{PORT}",
                width=1440, height=920,
                min_size=(1024, 700),
            )
            # Nhấn X = tắt sạch: dừng luôn runner đăng nền + join worker
            win.events.closing += _shutdown_all
            # block đến khi đóng cửa sổ → tiến trình kết thúc
            try:
                webview.start(icon=_icon)
            except TypeError:
                webview.start()   # pywebview cũ không hỗ trợ tham số icon
